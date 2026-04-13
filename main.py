import io
import os
import json
import sys
import traceback
import requests
from urllib.parse import quote
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Protection, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import Flask, request as flask_request, make_response

import hmac
import hashlib
import base64
import uuid

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

app = Flask(__name__)

# ── Firebase Admin SDK ────────────────────────────────────────────────────────
import firebase_admin
from firebase_admin import credentials, firestore as admin_firestore, auth as fb_auth

_fdb = None

def _get_db():
    """Initialise Firebase Admin once, return Firestore client."""
    global _fdb
    if _fdb:
        return _fdb
    if not firebase_admin._apps:
        sa_json = os.environ.get(\'SERVICE_ACCOUNT_JSON\', \'\')
        if sa_json:
            try:
                sa_dict = json.loads(sa_json)
            except json.JSONDecodeError:
                sa_dict = json.loads(sa_json.replace(\'\\r\\n\', \'\\n\'))
            if \'private_key\' in sa_dict:
                pk = sa_dict[\'private_key\']
                if \'\\\\n\' in pk:
                    sa_dict[\'private_key\'] = pk.replace(\'\\\\n\', \'\\n\')
            try:
                cred = credentials.Certificate(sa_dict)
            except Exception:
                traceback.print_exc()
                raise
        else:
            cred = credentials.Certificate(\'serviceAccountKey.json\')
        try:
            firebase_admin.initialize_app(cred)
        except Exception:
            traceback.print_exc()
            raise
    try:
        _fdb = admin_firestore.client()
        return _fdb
    except Exception:
        traceback.print_exc()
        raise


# ── Environment / constants ───────────────────────────────────────────────────
PROTECT_PASSWORD = os.environ.get(\'XLSX_PASSWORD\', \'EduTestPro2025\')
RDS_URL = os.environ.get(\'RDS_URL\', \'https://rds.edutest-pro.online\')

ALLOWED_CBT_ORIGINS = {
    \'https://edutest-pro.online\',
    \'https://edutest-pro-cbt.web.app\',
    \'https://edutest-pro-cbt.firebaseapp.com\',
}
ALLOWED_RDS_ORIGINS = {
    \'https://rds.edutest-pro.online\',
    \'https://edutest-rds.web.app\',
    \'https://edutest-rds.firebaseapp.com\',
}
ALL_ALLOWED_ORIGINS = ALLOWED_CBT_ORIGINS | ALLOWED_RDS_ORIGINS


def _cors_headers(req=None):
    origin  = (req.headers.get(\'Origin\', \'\') if req else \'\') or \'\'
    allowed = origin if origin in ALLOWED_CBT_ORIGINS else \'https://edutest-pro.online\'
    return {
        \'Access-Control-Allow-Origin\':  allowed,
        \'Access-Control-Allow-Methods\': \'POST, OPTIONS\',
        \'Access-Control-Allow-Headers\': \'Content-Type, Authorization, X-EduTest-Key\',
        \'Vary\': \'Origin\',
    }


def _make_rds_cors(req=None):
    origin     = (req.headers.get(\'Origin\', \'\') if req else \'\') or \'\'
    is_firebase = origin.endswith(\'.web.app\') or origin.endswith(\'.firebaseapp.com\')
    if origin in ALL_ALLOWED_ORIGINS or is_firebase:
        allowed = origin
    else:
        allowed = RDS_URL.rstrip(\'/\')
    return {
        \'Access-Control-Allow-Origin\':      allowed,
        \'Access-Control-Allow-Methods\':     \'POST, GET, OPTIONS\',
        \'Access-Control-Allow-Headers\':     \'Content-Type, Authorization, X-EduTest-Key\',
        \'Access-Control-Allow-Credentials\': \'false\',
        \'Vary\': \'Origin\',
    }

CORS = _cors_headers()


# ── Style constants ────────────────────────────────────────────────────────────
_GREEN       = \'1B6B45\'
_GREEN_LIGHT = \'2D9B6A\'
_DARK        = \'1A1A2E\'
_GREY        = \'6B7280\'
_WHITE       = \'FFFFFF\'
_PASS_CLR    = \'059669\'
_FAIL_CLR    = \'DC2626\'
_ALT_BG      = \'F0FDF4\'

_TITLE_FILL   = PatternFill(\'solid\', fgColor=_GREEN)
_SUB_FILL     = PatternFill(\'solid\', fgColor=_GREEN_LIGHT)
_COL_HDR_FILL = PatternFill(\'solid\', fgColor=\'166534\')
_ALT_FILL     = PatternFill(\'solid\', fgColor=_ALT_BG)
_LOCK         = Protection(locked=True)

def _font(size=10, bold=False, color=_DARK, name=\'Calibri\'):
    return Font(name=name, size=size, bold=bold, color=color)

_CENTER = Alignment(horizontal=\'center\', vertical=\'center\', wrap_text=False)
_LEFT   = Alignment(horizontal=\'left\',   vertical=\'center\')
_THIN   = Side(style=\'thin\',   color=\'D1FAE5\')
_MED    = Side(style=\'medium\', color=\'059669\')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

NUM_COLS = 9


# ── Shared helpers ─────────────────────────────────────────────────────────────

def _verify_token(request):
    _get_db()
    hdr = request.headers.get(\'Authorization\', \'\')
    if not hdr.startswith(\'Bearer \'):
        raise PermissionError(\'Missing or invalid Authorization header\')
    decoded = fb_auth.verify_id_token(hdr[7:])
    return decoded[\'email\'].lower(), decoded[\'uid\']


def _secret_ok(request):
    secret = os.environ.get(\'XLSX_SECRET\', \'\')
    if secret and request.headers.get(\'X-EduTest-Key\', \'\') != secret:
        raise PermissionError(\'Unauthorized\')


def _json_resp(data, status=200, req=None):
    resp = make_response(json.dumps(data), status)
    resp.headers[\'Content-Type\'] = \'application/json\'
    for k, v in _cors_headers(req).items():
        resp.headers[k] = v
    return resp


def _rds_json_resp(data, status=200, req=None):
    resp = make_response(json.dumps(data), status)
    resp.headers[\'Content-Type\'] = \'application/json\'
    for k, v in _make_rds_cors(req).items():
        resp.headers[k] = v
    return resp


def _err(msg, status=400, req=None):
    return _json_resp({\'ok\': False, \'error\': msg}, status, req)


def _rds_session(request, db):
    """Validate RDS session. Returns session dict or raises PermissionError."""
    session_id = (request.headers.get(\'Authorization\') or \'\').replace(\'Session \', \'\').strip()
    if not session_id:
        raise PermissionError(\'No session provided\')
    snap = db.collection(\'rds_sessions\').document(session_id).get()
    if not snap.exists:
        raise PermissionError(\'Invalid session\')
    sess = snap.to_dict()
    expires_at = sess.get(\'expiresAt\')
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at.replace(\'Z\', \'+00:00\'))
    if expires_at and datetime.now(timezone.utc) > expires_at:
        db.collection(\'rds_sessions\').document(session_id).delete()
        raise PermissionError(\'Session expired\')
    return sess


def _ts(dt_field):
    if dt_field and hasattr(dt_field, \'isoformat\'):
        return dt_field.isoformat()
    return dt_field


def _push_notification(db, recipient_uid, school_id, notif_type, title, body_text, meta=None):
    """Write an in-app notification to rds_notifications."""
    try:
        db.collection(\'rds_notifications\').add({
            \'recipientUid\': recipient_uid,
            \'schoolId\':     school_id,
            \'type\':         notif_type,
            \'title\':        title,
            \'body\':         body_text,
            \'meta\':         meta or {},
            \'read\':         False,
            \'createdAt\':    datetime.now(timezone.utc),
        })
    except Exception:
        pass  # Notifications are best-effort — never block the main action


# =============================================================================
# EXAM ENDPOINTS
# =============================================================================

@app.route(\'/get-exam\', methods=[\'POST\', \'OPTIONS\'])
def get_exam():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        email, _ = _verify_token(request)
        body     = request.get_json(force=True, silent=True) or {}
        exam_id  = (body.get(\'examId\') or \'\').strip()
        if not exam_id:
            return _err(\'examId required\')

        db = _get_db()
        u  = db.collection(\'users\').document(email).get()
        if not u.exists:
            return _err(\'Account not found\', 403)
        ud = u.to_dict()
        if ud.get(\'status\') != \'active\':
            return _err(\'Account suspended\', 403)
        if ud.get(\'role\') != \'student\':
            return _err(\'Only students can take exams\', 403)

        ex = db.collection(\'exams\').document(exam_id).get()
        if not ex.exists:
            return _err(\'Exam not found\', 404)
        ed = ex.to_dict()

        if ed.get(\'schoolId\') and ed[\'schoolId\'] != ud.get(\'schoolId\'):
            return _err(\'Exam not available for your school\', 403)

        safe_qs = [
            {\'question\': q.get(\'question\', \'\'), \'options\': q.get(\'options\', [])}
            for q in ed.get(\'questions\', [])
        ]

        return _json_resp({\'ok\': True, \'exam\': {
            \'id\':               exam_id,
            \'title\':            ed.get(\'title\', \'\'),
            \'description\':      ed.get(\'description\', \'\'),
            \'duration_minutes\': ed.get(\'duration_minutes\', 60),
            \'schoolId\':         ed.get(\'schoolId\', \'\'),
            \'questions\':        safe_qs,
        }})

    except PermissionError as e:
        return _err(str(e), 403)
    except Exception as e:
        return _err(str(e), 500)


@app.route(\'/check-submitted\', methods=[\'POST\', \'OPTIONS\'])
def check_submitted():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        email, _ = _verify_token(request)
        body    = request.get_json(force=True, silent=True) or {}
        exam_id = (body.get(\'examId\') or \'\').strip()
        if not exam_id:
            return _err(\'examId required\')
        snap = _get_db().collection(\'submissions\').document(exam_id + \'_\' + email).get()
        return _json_resp({\'ok\': True, \'submitted\': snap.exists})
    except PermissionError as e:
        return _err(str(e), 403)
    except Exception as e:
        return _err(str(e), 500)


@app.route(\'/submit-exam\', methods=[\'POST\', \'OPTIONS\'])
def submit_exam():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        email, _       = _verify_token(request)
        body           = request.get_json(force=True, silent=True) or {}
        exam_id        = (body.get(\'examId\') or \'\').strip()
        raw_answers    = body.get(\'rawAnswers\', {})
        question_order = body.get(\'questionOrder\', [])
        option_orders  = body.get(\'optionOrders\', [])
        time_taken     = int(body.get(\'timeTaken\', 0))

        if not exam_id:
            return _err(\'examId required\')
        if not isinstance(raw_answers, dict) or len(raw_answers) > 500:
            return _err(\'Invalid or oversized answer payload\', 400)
        if not isinstance(question_order, list) or len(question_order) > 500:
            return _err(\'Invalid questionOrder\', 400)
        if not isinstance(option_orders, list) or len(option_orders) > 500:
            return _err(\'Invalid optionOrders\', 400)

        db = _get_db()
        u  = db.collection(\'users\').document(email).get()
        if not u.exists:
            return _err(\'Account not found\', 403)
        ud = u.to_dict()
        if ud.get(\'status\') != \'active\' or ud.get(\'role\') != \'student\':
            return _err(\'Not authorised\', 403)

        doc_id = exam_id + \'_\' + email
        if db.collection(\'submissions\').document(doc_id).get().exists:
            return _err(\'You have already submitted this exam.\', 409)

        ex = db.collection(\'exams\').document(exam_id).get()
        if not ex.exists:
            return _err(\'Exam not found\', 404)
        ed = ex.to_dict()

        if ed.get(\'schoolId\') and ed[\'schoolId\'] != ud.get(\'schoolId\'):
            return _err(\'Exam not available for your school\', 403)

        close_date_str = ed.get(\'closeDate\')
        if close_date_str:
            try:
                close_dt = datetime.fromisoformat(close_date_str.replace(\'Z\', \'+00:00\'))
                if datetime.now(timezone.utc) > close_dt:
                    return _err(\'This exam has closed and is no longer available.\', 403)
            except Exception:
                pass

        orig_qs = ed.get(\'questions\', [])
        total   = len(orig_qs)

        duration_minutes = ed.get(\'duration_minutes\', 0)
        if duration_minutes and time_taken > 0:
            max_allowed_secs = duration_minutes * 60 + 60
            if time_taken > max_allowed_secs:
                time_taken = max_allowed_secs

        if not question_order:
            question_order = list(range(total))
        if not option_orders:
            option_orders = [list(range(len(q.get(\'options\', [])))) for q in orig_qs]

        correct  = 0
        answered = 0
        audit    = {}

        for si, orig_qi in enumerate(question_order):
            if orig_qi >= len(orig_qs):
                continue
            oq        = orig_qs[orig_qi]
            opt_order = option_orders[si] if si < len(option_orders) else list(range(len(oq.get(\'options\', []))))
            correct_idx = oq.get(\'correctIndex\', -1)
            options     = oq.get(\'options\', [])

            picked_s = raw_answers.get(str(si))
            if picked_s is not None:
                picked_s    = int(picked_s)
                answered   += 1
                picked_orig = opt_order[picked_s] if picked_s < len(opt_order) else None
            else:
                picked_orig = None

            is_correct = (picked_orig is not None and picked_orig == correct_idx)
            if is_correct:
                correct += 1

            audit[str(orig_qi)] = {
                \'questionText\': oq.get(\'question\', \'\'),
                \'pickedText\':   options[picked_orig] if picked_orig is not None and picked_orig < len(options) else \'(not answered)\',
                \'correctText\':  options[correct_idx] if 0 <= correct_idx < len(options) else \'\',
                \'isCorrect\':    is_correct,
                \'notAnswered\':  picked_orig is None,
            }

        wrong      = answered - correct
        unanswered = total - answered
        percentage = round(correct / total * 100) if total else 0

        db.collection(\'submissions\').document(doc_id).set({
            \'examId\':       exam_id,
            \'examTitle\':    ed.get(\'title\', \'\'),
            \'schoolId\':     ed.get(\'schoolId\', \'\'),
            \'studentEmail\': email,
            \'studentName\':  (ud.get(\'name\') or \'\').strip(),
            \'studentClass\': (ud.get(\'classGrade\') or \'\').strip(),
            \'answers\':      audit,
            \'score\':        correct,
            \'total\':        total,
            \'wrong\':        wrong,
            \'unanswered\':   unanswered,
            \'percentage\':   percentage,
            \'timeTaken\':    time_taken,
            'submittedAt':  admin_firestore.FieldValue.server_timestamp(),
        })

        return _json_resp({
            \'ok\': True, \'correct\': correct, \'wrong\': wrong,
            \'unanswered\': unanswered, \'total\': total, \'percentage\': percentage,
        })

    except PermissionError as e:
        return _err(str(e), 403)
    except Exception as e:
        traceback.print_exc()
        return _err(str(e), 500)


@app.route(\'/list-exams\', methods=[\'POST\', \'OPTIONS\'])
def list_exams():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        email, _ = _verify_token(request)
        db = _get_db()

        u = db.collection(\'users\').document(email).get()
        if not u.exists:
            return _err(\'Account not found\', 403)
        ud = u.to_dict()
        if ud.get(\'status\') != \'active\':
            return _err(\'Account suspended\', 403)
        if ud.get(\'role\') != \'student\':
            return _err(\'Only students can list exams\', 403)

        school_id = ud.get(\'schoolId\', \'\')
        if not school_id:
            return _err(\'No school assigned to this account\', 403)

        exams_snap = db.collection(\'exams\').where(\'schoolId\', \'==\', school_id).get()
        subs_snap  = (db.collection(\'submissions\')
                       .where(\'studentEmail\', \'==\', email.lower())
                       .where(\'schoolId\',     \'==\', school_id)
                       .get())
        submitted_ids = {s.to_dict().get(\'examId\') for s in subs_snap}

        now       = datetime.now(timezone.utc)
        exams_out = []
        for ex in exams_snap:
            ed = ex.to_dict()
            if ex.id in submitted_ids:
                continue
            close_date_str = ed.get(\'closeDate\')
            if close_date_str:
                try:
                    close_dt = datetime.fromisoformat(close_date_str.replace(\'Z\', \'+00:00\'))
                    if now > close_dt:
                        continue
                except Exception:
                    pass
            exams_out.append({
                \'id\':               ex.id,
                \'title\':            ed.get(\'title\', \'\'),
                \'description\':      ed.get(\'description\', \'\'),
                \'duration_minutes\': ed.get(\'duration_minutes\', 60),
                \'targetClass\':      ed.get(\'targetClass\', \'\'),
                \'examTerm\':         ed.get(\'examTerm\', \'\'),
                \'examType\':         ed.get(\'examType\', \'\'),
                \'questionCount\':    len(ed.get(\'questions\', [])),
                \'closeDate\':        close_date_str or None,
            })

        exams_out.sort(key=lambda e: e[\'title\'])
        return _json_resp({\'ok\': True, \'exams\': exams_out})

    except PermissionError as e:
        traceback.print_exc()
        return _err(str(e), 403)
    except Exception as e:
        traceback.print_exc()
        return _err(str(e), 500)


# =============================================================================
# XLSX ENDPOINTS
# =============================================================================

@app.route(\'/generate_results_xlsx\', methods=[\'POST\', \'OPTIONS\'])
def generate_results_xlsx():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))

    api_secret = os.environ.get(\'XLSX_SECRET\', \'\')
    if api_secret and request.headers.get(\'X-EduTest-Key\', \'\') != api_secret:
        return ('Unauthorized', 401, _make_rds_cors(request))
    try:
        caller_email, _ = _verify_token(request)
    except PermissionError as e:
        return _err(str(e), 401)

    db = _get_db()
    caller_doc = db.collection(\'users\').document(caller_email).get()
    if not caller_doc.exists:
        return _err(\'Account not found\', 403)
    if caller_doc.to_dict().get(\'role\') not in (\'super_admin\', \'school_admin\', \'sub_admin\', \'teacher\'):
        return _err(\'Not authorised\', 403)

    body             = request.get_json(force=True, silent=True) or {}
    title            = str(body.get(\'title\',            \'Exam Results\'))[:120]
    school           = str(body.get(\'school\',           \'\'))[:120]
    academic_session = str(body.get(\'academic_session\', \'\'))[:80]
    term             = str(body.get(\'term\',             \'\'))[:60]
    exam_type        = str(body.get(\'exam_type\',        \'\'))[:60]
    downloaded_by    = str(body.get(\'downloaded_by\',    \'\'))[:120]
    rows             = body.get(\'rows\', [])

    data     = _build_xlsx(title, school, academic_session, term, exam_type, downloaded_by, rows)
    safe     = \'\'.join(c if c.isalnum() or c in \' _-\' else \'_\' for c in title)
    filename = f"{safe}.xlsx"

    resp = make_response(data)
    resp.headers[\'Content-Type\']        = \'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\'
    resp.headers[\'Content-Disposition\'] = f\'attachment; filename="{filename}"\'
    for k, v in _make_rds_cors(request).items():
        resp.headers[k] = v
    return resp


def _lock_all(ws):
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = _LOCK
    ws.protection.set_password(PROTECT_PASSWORD)
    ws.protection.sheet               = True
    ws.protection.insertRows          = True
    ws.protection.insertColumns       = True
    ws.protection.deleteRows          = True
    ws.protection.deleteColumns       = True
    ws.protection.formatCells         = True
    ws.protection.formatRows          = True
    ws.protection.formatColumns       = True
    ws.protection.sort                = True
    ws.protection.autoFilter          = True
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False


def _merged(ws, row, col_start, col_end, value, fill, font, align=None):
    if col_start < col_end:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.fill      = fill
    c.font      = font
    c.alignment = align or _CENTER
    return c


def _build_xlsx(title, school, academic_session, term, exam_type, downloaded_by, rows):
    wb = Workbook()
    ws = wb.active
    ws.title        = \'Results\'
    ws.freeze_panes = \'A6\'

    _merged(ws, 1, 1, NUM_COLS,
            school.upper() if school else \'SCHOOL RESULTS\',
            _TITLE_FILL, _font(14, True, _WHITE, \'Calibri\'))
    ws.row_dimensions[1].height = 32

    _merged(ws, 2, 1, 3,
            f\'Academic Session: {academic_session}\' if academic_session else \'Academic Session: —\',
            _SUB_FILL, _font(10, True, _WHITE))
    _merged(ws, 2, 4, 6,
            f\'Term: {term}\' if term else \'Term: —\',
            _SUB_FILL, _font(10, True, _WHITE))
    _merged(ws, 2, 7, NUM_COLS,
            f\'Type: {exam_type}\' if exam_type else \'Type: —\',
            _SUB_FILL, _font(10, True, _WHITE))
    ws.row_dimensions[2].height = 22

    _merged(ws, 3, 1, NUM_COLS, title,
            PatternFill(\'solid\', fgColor=\'F0FDF4\'), _font(12, True, _GREEN))
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 18

    headers = [\'#\', \'Full Name\', \'Email Address\', \'Class\', \'Arm\',
               \'Score\', \'Percentage\', \'Result\', \'Date Submitted\']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill      = _COL_HDR_FILL
        c.font      = _font(10, True, _WHITE)
        c.alignment = _CENTER
        c.border    = Border(
            left=Side(style=\'medium\', color=_WHITE),
            right=Side(style=\'medium\', color=_WHITE),
            bottom=Side(style=\'medium\', color=_WHITE))
    ws.row_dimensions[5].height = 22

    pass_count = 0
    pct_total  = 0

    for i, row in enumerate(rows):
        r    = i + 6
        pct  = int(row.get(\'percentage\', 0))
        rslt = \'PASS\' if pct >= 50 else \'FAIL\'
        alt  = (i % 2 == 0)
        if rslt == \'PASS\': pass_count += 1
        pct_total += pct

        values = [
            i + 1,
            row.get(\'name\',  \'—\'),
            row.get(\'email\', \'\'),
            row.get(\'class\', \'—\'),
            row.get(\'arm\',   \'—\'),
            f"{row.get(\'score\',0)}/{row.get(\'total\',0)}",
            f"{pct}%",
            rslt,
            row.get(\'submitted\', \'—\'),
        ]
        aligns = [_CENTER, _LEFT, _LEFT, _CENTER, _CENTER,
                  _CENTER, _CENTER, _CENTER, _CENTER]
        colors_list = [_GREY, _DARK, _GREY, _DARK, _DARK,
                  _DARK, _DARK,
                  _PASS_CLR if rslt == \'PASS\' else _FAIL_CLR,
                  _GREY]
        bolds  = [False, True, False, False, False,
                  False, True, True, False]

        for col, (val, aln, clr, bld) in enumerate(zip(values, aligns, colors_list, bolds), 1):
            c           = ws.cell(row=r, column=col, value=val)
            c.alignment = aln
            c.border    = _BORDER
            c.font      = _font(9 if col in (3, 9) else 10, bld, clr)
            if alt:
                c.fill = _ALT_FILL
        ws.row_dimensions[r].height = 18

    total      = len(rows)
    avg        = round(pct_total / total) if total else 0
    fail_count = total - pass_count
    pass_rate  = round(pass_count / total * 100) if total else 0

    stats     = [
        (\'Total\', str(total)),
        (\'Passed\', str(pass_count)),
        (\'Failed\', str(fail_count)),
        (\'Average\', f\'{avg}%\'),
        (\'Pass Rate\', f\'{pass_rate}%\'),
    ]
    stat_fill = PatternFill(\'solid\', fgColor=\'DCFCE7\')
    for idx, (lbl, val) in enumerate(stats):
        lc = idx * 2 + 1
        if lc + 1 > NUM_COLS:
            break
        l_cell = ws.cell(row=4, column=lc, value=f\'{lbl}: {val}\')
        l_cell.font      = _font(9, True, _GREEN)
        l_cell.alignment = _CENTER
        l_cell.fill      = stat_fill

    widths = [5, 28, 32, 10, 6, 10, 13, 10, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _lock_all(ws)

    ws2     = wb.create_sheet(\'Info\')
    now_str = datetime.now(timezone.utc).strftime(\'%d %b %Y  %H:%M UTC\')
    info_rows = [
        (\'EduTest Pro — Results Export\', \'\'),
        (\'\', \'\'),
        (\'School\',           school),
        (\'Academic Session\', academic_session or \'—\'),
        (\'Term\',             term             or \'—\'),
        (\'Exam Type\',        exam_type        or \'—\'),
        (\'Report Title\',     title),
        (\'Downloaded By\',    downloaded_by),
        (\'Download Date\',    now_str),
        (\'Total Records\',    total),
        (\'\', \'\'),
        (\'Protection Note\',  \'This file is read-only and protected by EduTest Pro.\'),
        (\'Unlock Password\',  \'Contact your system administrator for the password.\'),
    ]
    for r, (k, v) in enumerate(info_rows, 1):
        ck = ws2.cell(row=r, column=1, value=k)
        cv = ws2.cell(row=r, column=2, value=v)
        ck.font = _font(10, True,  _GREY)
        cv.font = _font(10, False, _DARK)
        if r == 1:
            cv.value = \'\'
            ws2.merge_cells(\'A1:B1\')
            ck.font = _font(13, True, _GREEN)
    ws2.column_dimensions[\'A\'].width = 22
    ws2.column_dimensions[\'B\'].width = 58
    _lock_all(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# ANSWER AUDIT ENDPOINT
# =============================================================================

AUDIT_NCOLS = 5


@app.route(\'/generate_audit_xlsx\', methods=[\'POST\', \'OPTIONS\'])
def generate_audit_xlsx():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))

    api_secret = os.environ.get(\'XLSX_SECRET\', \'\')
    if api_secret and request.headers.get(\'X-EduTest-Key\', \'\') != api_secret:
        return ('Unauthorized', 401, _make_rds_cors(request))
    try:
        caller_email, _ = _verify_token(request)
    except PermissionError as e:
        return _err(str(e), 401)

    db = _get_db()
    caller_doc = db.collection(\'users\').document(caller_email).get()
    if not caller_doc.exists:
        return _err(\'Account not found\', 403)
    if caller_doc.to_dict().get(\'role\') not in (\'super_admin\', \'school_admin\', \'sub_admin\', \'teacher\'):
        return _err(\'Not authorised\', 403)

    body             = request.get_json(force=True, silent=True) or {}
    title            = str(body.get(\'title\',            \'Answer Audit\'))[:120]
    school           = str(body.get(\'school\',           \'\'))[:120]
    academic_session = str(body.get(\'academic_session\', \'\'))[:80]
    term             = str(body.get(\'term\',             \'\'))[:60]
    exam_type        = str(body.get(\'exam_type\',        \'\'))[:60]
    requested_by     = str(body.get(\'requested_by\',     \'\'))[:120]
    scope            = str(body.get(\'scope\',            \'\'))[:120]
    rows             = body.get(\'rows\', [])

    data     = _build_audit_xlsx(title, school, academic_session, term,
                                  exam_type, requested_by, scope, rows)
    safe     = \'\'.join(ch if ch.isalnum() or ch in \' _-\' else \'_\' for ch in title)
    filename = f"{safe}.xlsx"

    resp = make_response(data)
    resp.headers[\'Content-Type\']        = \'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\'
    resp.headers[\'Content-Disposition\'] = f\'attachment; filename="{filename}"\'
    for k, v in _make_rds_cors(request).items():
        resp.headers[k] = v
    return resp


def _build_audit_xlsx(title, school, academic_session, term, exam_type,
                       requested_by, scope, rows):
    AMBER        = \'B45309\'
    AMBER_LIGHT  = \'D97706\'
    AMBER_PALE   = \'FEF3C7\'
    AMBER_BG     = \'FFFBEB\'
    GREEN_OK     = \'059669\'
    RED_NO       = \'DC2626\'
    GREY_NA      = \'9CA3AF\'
    STUDENT_BG   = \'166534\'

    FILL_AMBER       = PatternFill(\'solid\', fgColor=AMBER)
    FILL_AMBER_LIGHT = PatternFill(\'solid\', fgColor=AMBER_LIGHT)
    FILL_AMBER_PALE  = PatternFill(\'solid\', fgColor=AMBER_PALE)
    FILL_AMBER_BG    = PatternFill(\'solid\', fgColor=AMBER_BG)
    FILL_WARN        = PatternFill(\'solid\', fgColor=\'FEE2E2\')
    FILL_STUDENT     = PatternFill(\'solid\', fgColor=STUDENT_BG)
    FILL_META        = PatternFill(\'solid\', fgColor=\'F0FDF4\')
    FILL_COL_HDR     = PatternFill(\'solid\', fgColor=\'92400E\')

    NC = AUDIT_NCOLS
    wb = Workbook()
    ws = wb.active
    ws.title = \'Answer Audit\'
    row = 1

    def cell_m(r, c1, c2, val, fill, fnt, aln=None):
        if c1 < c2:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cel = ws.cell(row=r, column=c1, value=val)
        cel.fill = fill; cel.font = fnt; cel.alignment = aln or _CENTER
        return cel

    cell_m(row, 1, NC, school.upper() if school else \'ANSWER AUDIT\',
           FILL_AMBER, _font(13, True, _WHITE))
    ws.row_dimensions[row].height = 30; row += 1

    cell_m(row, 1, 2,
           f\'Session: {academic_session}\' if academic_session else \'Session: —\',
           FILL_AMBER_LIGHT, _font(9, True, _WHITE))
    ws.cell(row=row, column=3, value=f\'Term: {term}\' if term else \'Term: —\').fill = FILL_AMBER_LIGHT
    ws.cell(row=row, column=3).font      = _font(9, True, _WHITE)
    ws.cell(row=row, column=3).alignment = _CENTER
    cell_m(row, 4, NC, f\'Type: {exam_type}\' if exam_type else \'Type: —\',
           FILL_AMBER_LIGHT, _font(9, True, _WHITE))
    ws.row_dimensions[row].height = 18; row += 1

    cell_m(row, 1, NC, title, FILL_AMBER_PALE, _font(11, True, AMBER))
    ws.row_dimensions[row].height = 22; row += 1

    cell_m(row, 1, NC,
           \'⚠  CONFIDENTIAL — Authorised Investigation Use Only\',
           FILL_WARN, _font(9, True, RED_NO))
    ws.row_dimensions[row].height = 16; row += 1

    cell_m(row, 1, 2, f\'Scope: {scope}\' if scope else \'\',
           FILL_AMBER_PALE, _font(9, False, AMBER))
    cell_m(row, 3, NC, f\'Requested by: {requested_by}\',
           FILL_AMBER_PALE, _font(9, False, AMBER))
    ws.row_dimensions[row].height = 14; row += 1
    row += 1  # spacer

    for stu in rows:
        q_rows     = stu.get(\'q_rows\', [])
        pct        = int(stu.get(\'percentage\', 0))
        score      = stu.get(\'score\', 0)
        total_qs   = stu.get(\'total_qs\', 0)
        answered   = stu.get(\'answered\', 0)
        unanswered = stu.get(\'unanswered\', 0)

        cell_m(row, 1, NC,
               f\'{stu.get("name","—")}   ·   {stu.get("email","")}\',
               FILL_STUDENT, _font(10, True, _WHITE))
        ws.row_dimensions[row].height = 20; row += 1

        meta = [
            f\'Class: {stu.get("class","—")}\',
            f\'Exam: {stu.get("exam","—")}\',
            f\'Score: {score}/{total_qs}  ({pct}%)\',
            f\'Answered: {answered}  ·  Unanswered: {unanswered}\',
            f\'Submitted: {stu.get("submitted","—")}\',
        ]
        for col, val in enumerate(meta, 1):
            cel = ws.cell(row=row, column=col, value=val)
            cel.fill = FILL_META; cel.font = _font(9, False, _DARK); cel.alignment = _LEFT
        ws.row_dimensions[row].height = 16; row += 1

        for col, hdr in enumerate([\'Q#\', \'Question Text\', \'Option Picked\', \'Correct Answer\', \'Correct?\'], 1):
            cel = ws.cell(row=row, column=col, value=hdr)
            cel.fill = FILL_COL_HDR; cel.font = _font(9, True, _WHITE); cel.alignment = _CENTER
        ws.row_dimensions[row].height = 18; row += 1

        if not q_rows:
            cell_m(row, 1, NC, \'(no answer data recorded)\', PatternFill(), _font(9, False, GREY_NA))
            ws.row_dimensions[row].height = 16; row += 1
        else:
            for qi, qr in enumerate(q_rows):
                status       = qr.get(\'is_correct\', \'NO\')
                is_yes       = status == \'YES\'
                is_na        = status == \'NOT ANSWERED\'
                alt_fill     = FILL_AMBER_BG if qi % 2 == 0 else PatternFill()
                picked_color = GREEN_OK if is_yes else (GREY_NA if is_na else RED_NO)
                status_color = GREEN_OK if is_yes else (GREY_NA if is_na else RED_NO)

                vals   = [qr.get(\'q_num\', qi+1), qr.get(\'question\',\'—\'),
                          qr.get(\'picked\',\'(not answered)\'), qr.get(\'correct\',\'—\'), status]
                aligns = [_CENTER, _LEFT, _LEFT, _LEFT, _CENTER]
                clrs   = [_GREY, _DARK, picked_color, GREEN_OK, status_color]
                bolds  = [False,  False, True,         True,     True]

                for col, (v, aln, clr, bld) in enumerate(zip(vals, aligns, clrs, bolds), 1):
                    cel = ws.cell(row=row, column=col, value=v)
                    cel.border = _BORDER; cel.font = _font(9, bld, clr); cel.alignment = aln
                    if qi % 2 == 0: cel.fill = alt_fill

                for col in [2, 3, 4]:
                    ws.cell(row=row, column=col).alignment = Alignment(
                        horizontal=\'left\', vertical=\'top\', wrap_text=True)
                ws.row_dimensions[row].height = 32; row += 1

        ws.row_dimensions[row].height = 6; row += 1

    for col, width in enumerate([4, 52, 32, 32, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _lock_all(ws)

    ws2 = wb.create_sheet(\'Audit Trail\')
    now = datetime.now(timezone.utc).strftime(\'%d %b %Y  %H:%M UTC\')
    trail = [
        (\'EduTest Pro — Answer Audit\', \'\'),
        (\'\', \'\'),
        (\'School\',           school),
        (\'Academic Session\', academic_session or \'—\'),
        (\'Term\',             term             or \'—\'),
        (\'Exam Type\',        exam_type        or \'—\'),
        (\'Report Title\',     title),
        (\'Scope\',            scope),
        (\'Requested By\',     requested_by),
        (\'Generated On\',     now),
        (\'Total Students\',   len(rows)),
        (\'\', \'\'),
        (\'Contents\',         \'Question text · Option the student selected · Correct answer · Whether correct.\'),
        (\'Does NOT contain\', \'All answer options are NOT listed — only what was picked and what was correct.\'),
        (\'Intended use\',     \'Authorised investigation only. Protected by EduTest Pro.\'),
    ]
    for r, (k, v) in enumerate(trail, 1):
        ck = ws2.cell(row=r, column=1, value=k)
        cv = ws2.cell(row=r, column=2, value=v)
        ck.font = _font(10, True, _GREY); cv.font = _font(10, False, _DARK)
        if r == 1:
            ws2.merge_cells(\'A1:B1\'); cv.value = \'\'
            ck.font = _font(13, True, AMBER)
    ws2.column_dimensions[\'A\'].width = 20
    ws2.column_dimensions[\'B\'].width = 80
    _lock_all(ws2)

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# =============================================================================
# PAYSTACK PAYMENT ENDPOINTS
# =============================================================================

@app.route(\'/init-payment\', methods=[\'POST\', \'OPTIONS\'])
def init_payment():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        caller_email, _ = _verify_token(request)

        db = _get_db()
        caller_doc = db.collection(\'users\').document(caller_email).get()
        if not caller_doc.exists:
            return _err(\'Account not found\', 403, request)
        caller_data = caller_doc.to_dict()
        if caller_data.get(\'role\') not in (\'school_admin\', \'super_admin\'):
            return _err(\'Not authorised\', 403, request)

        body      = request.get_json(force=True, silent=True) or {}
        school_id = (body.get(\'schoolId\')    or \'\').strip()
        amount    = int(body.get(\'amount\', 0))
        email     = (body.get(\'email\')       or \'\').strip()
        reference = (body.get(\'reference\')   or \'\').strip()
        callback  = (body.get(\'callbackUrl\') or \'https://edutest-pro.online/app.html\').strip()

        if not school_id or not amount or not email:
            return _err(\'schoolId, amount, and email are required\', 400, request)

        if caller_data.get(\'role\') == \'school_admin\':
            if caller_data.get(\'schoolId\') != school_id:
                return _err(\'Not authorised for this school\', 403, request)

        paystack_secret = os.environ.get(\'PAYSTACK_SECRET_KEY\', \'\')
        if not paystack_secret:
            return _err(\'Payment gateway not configured on server\', 500, request)

        resp = requests.post(
            \'https://api.paystack.co/transaction/initialize\',
            headers={
                \'Authorization\': f\'Bearer {paystack_secret}\',
                \'Content-Type\':  \'application/json\',
            },
            json={
                \'email\':        email,
                \'amount\':       amount * 100,
                \'currency\':     \'NGN\',
                \'reference\':    reference,
                \'callback_url\': callback,
                \'metadata\': {
                    \'custom_fields\': [
                        {\'display_name\': \'School ID\',  \'variable_name\': \'school_id\',  \'value\': school_id},
                        {\'display_name\': \'Renewed By\', \'variable_name\': \'renewed_by\', \'value\': email},
                    ],
                },
            },
            timeout=15,
        )
        data = resp.json()

        if not resp.ok or not data.get(\'status\') or not data.get(\'data\', {}).get(\'authorization_url\'):
            msg = data.get(\'message\', \'Paystack initialization failed\')
            return _err(msg, 400, request)

        return _json_resp({
            \'ok\':                True,
            \'authorization_url\': data[\'data\'][\'authorization_url\'],
            \'reference\':         data[\'data\'][\'reference\'],
            \'access_code\':       data[\'data\'].get(\'access_code\', \'\'),
        }, req=request)

    except PermissionError as e:
        return _err(str(e), 403, request)
    except Exception as e:
        traceback.print_exc()
        return _err(str(e), 500, request)


@app.route(\'/verify-payment\', methods=[\'POST\', \'OPTIONS\'])
def verify_payment():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _cors_headers(request))
    try:
        _secret_ok(request)
        caller_email, _ = _verify_token(request)

        db = _get_db()
        caller_doc = db.collection(\'users\').document(caller_email).get()
        if not caller_doc.exists:
            return _err(\'Account not found\', 403, request)
        caller_data = caller_doc.to_dict()
        if caller_data.get(\'role\') not in (\'school_admin\', \'super_admin\'):
            return _err(\'Not authorised\', 403, request)

        body      = request.get_json(force=True, silent=True) or {}
        reference = (body.get(\'reference\') or \'\').strip()
        school_id = (body.get(\'schoolId\')  or \'\').strip()
        amount    = int(body.get(\'amount\', 0))

        if not reference or not school_id:
            return _err(\'reference and schoolId required\', 400, request)

        paystack_secret = os.environ.get(\'PAYSTACK_SECRET_KEY\', \'\')
        if not paystack_secret:
            return _err(\'Payment gateway not configured\', 500, request)

        resp = requests.get(
            f\'https://api.paystack.co/transaction/verify/{reference}\',
            headers={\'Authorization\': f\'Bearer {paystack_secret}\'},
            timeout=10,
        )
        data = resp.json()

        if not data.get(\'status\') or data.get(\'data\', {}).get(\'status\') != \'success\':
            return _err(\'Payment not successful\', 402, request)

        paid_amount_kobo = data[\'data\'].get(\'amount\', 0)
        if paid_amount_kobo < amount * 100:
            return _err(\'Payment amount insufficient\', 402, request)

        school_snap = db.collection(\'schools\').document(school_id).get()
        if not school_snap.exists:
            return _err(\'School not found\', 404, request)
        school_data = school_snap.to_dict()

        if caller_data.get(\'role\') == \'school_admin\':
            if caller_data.get(\'schoolId\') != school_id:
                return _err(\'Not authorised for this school\', 403, request)

        sub_days = school_data.get(\'subscriptionDays\', 270)
        db.collection(\'schools\').document(school_id).update({
            \'subscriptionStart\':  datetime.now(timezone.utc).isoformat(),
            \'subscriptionDays\':   sub_days,
            \'lastPaymentRef\':     reference,
            \'lastPaymentAmount\':  paid_amount_kobo // 100,
            \'lastPaymentDate\':    datetime.now(timezone.utc).isoformat(),
        })

        return _json_resp({\'ok\': True, \'message\': \'Subscription activated\'}, req=request)

    except PermissionError as e:
        return _err(str(e), 403, request)
    except Exception as e:
        traceback.print_exc()
        return _err(str(e), 500, request)


# =============================================================================
# RDS — Token Bridge
# =============================================================================

def _rds_secret():
    secret = os.environ.get(\'RDS_BRIDGE_SECRET\', \'\')
    if not secret:
        raise PermissionError(\'RDS_BRIDGE_SECRET not configured on server\')
    return secret


@app.route(\'/launch-rds\', methods=[\'POST\', \'OPTIONS\'])
def launch_rds():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _make_rds_cors(request))
    try:
        _secret_ok(request)
        caller_email, _ = _verify_token(request)

        db = _get_db()
        caller_doc = db.collection(\'users\').document(caller_email).get()
        if not caller_doc.exists:
            return _rds_json_resp({\'ok\': False, \'error\': \'Account not found\'}, 403, request)

        caller = caller_doc.to_dict()
        if caller.get(\'role\') not in (\'super_admin\', \'school_admin\', \'sub_admin\', \'teacher\'):
            return _rds_json_resp({\'ok\': False,
                \'error\': \'Your role does not have access to Result Distribution System.\'}, 403, request)

        nonce  = uuid.uuid4().hex
        now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
        payload = {
            \'uid\':      caller_email,
            \'role\':     caller.get(\'role\'),
            \'schoolId\': caller.get(\'schoolId\', \'\'),
            \'email\':    caller_email,
            \'name\':     caller.get(\'name\', caller_email),
            \'iat\':      now_ms,
            \'exp\':      now_ms + 60000,
            \'nonce\':    nonce,
        }

        encoded = base64.urlsafe_b64encode(
            json.dumps(payload, separators=(\',\', \':\')).encode()
        ).rstrip(b\'=\').decode()

        sig = hmac.new(
            _rds_secret().encode(),
            encoded.encode(),
            hashlib.sha256
        ).hexdigest()

        token = f\'{encoded}.{sig}\'

        db.collection(\'rds_nonces\').document(nonce).set({
            \'uid\':       caller_email,
            \'schoolId\':  caller.get(\'schoolId\', \'\'),
            \'usedAt\':    None,
            \'expiresAt\': datetime.fromtimestamp((now_ms + 60000) / 1000, tz=timezone.utc),
            \'createdAt\': datetime.now(timezone.utc),
        })

        redirect_url = f'{RDS_URL.rstrip("/")}/access?token={quote(token)}'
        return _rds_json_resp({\'ok\': True, \'redirectUrl\': redirect_url}, req=request)

    except PermissionError as e:
        return _rds_json_resp({\'ok\': False, \'error\': str(e)}, 403, request)
    except Exception as e:
        traceback.print_exc()
        return _rds_json_resp({\'ok\': False, \'error\': str(e)}, 500, request)


@app.route(\'/rds-verify\', methods=[\'POST\', \'OPTIONS\'])
def rds_verify():
    request = flask_request
    if request.method == \'OPTIONS\':
        return make_response(\'\', 204, _make_rds_cors(request))
    try:
        body  = request.get_json(force=True, silent=True) or {}
        token = (body.get(\'token\') or \'\').strip()

        if not token or \'.\' not in token:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INVALID_TOKEN_FORMAT\',
                \'message\': \'Malformed access token. Please try again from CBT.\'}, 400, request)

        parts = token.rsplit(\'.\', 1)
        if len(parts) != 2:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INVALID_TOKEN_FORMAT\',
                \'message\': \'Malformed access token.\'}, 400, request)

        encoded, received_sig = parts

        expected_sig = hmac.new(
            _rds_secret().encode(),
            encoded.encode(),
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected_sig, received_sig):
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INVALID_SIGNATURE\',
                \'message\': \'Access token signature is invalid. Please try again from CBT.\'}, 401, request)

        try:
            padding = 4 - len(encoded) % 4
            padded  = encoded + \'=\' * (padding % 4)
            payload = json.loads(base64.urlsafe_b64decode(padded).decode())
        except Exception:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INVALID_PAYLOAD\',
                \'message\': \'Access token data is corrupted. Please try again from CBT.\'}, 400, request)

        required = (\'uid\', \'role\', \'email\', \'name\', \'iat\', \'exp\', \'nonce\')
        if not all(k in payload for k in required):
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INCOMPLETE_PAYLOAD\',
                \'message\': \'Access token is missing required data. Please try again from CBT.\'}, 400, request)

        now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
        if now_ms > payload[\'exp\']:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'TOKEN_EXPIRED\',
                \'message\': \'Access token has expired. Please click Result Distribution again in CBT.\'}, 401, request)

        nonce = payload.get(\'nonce\')
        if not nonce:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'MISSING_NONCE\',
                \'message\': \'Access token is missing security nonce.\'}, 400, request)

        db = _get_db()
        nonce_ref = db.collection(\'rds_nonces\').document(nonce)
        nonce_snap = nonce_ref.get()

        if not nonce_snap.exists:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'INVALID_NONCE\',
                \'message\': \'This token is invalid or has already been used.\'}, 401, request)

        nonce_data = nonce_snap.to_dict()
        if nonce_data.get(\'usedAt\') is not None:
            return _rds_json_resp({\'ok\': False, \'errorCode\': \'NONCE_ALREADY_USED\',
                \'message\': \'This token has already been consumed.\'}, 401, request)

        # Consume the nonce to prevent reuse
        nonce_ref.update({'usedAt': admin_firestore.FieldValue.server_timestamp()})

        # Create a persistent RDS session
        session_id = uuid.uuid4().hex
        expires_at = datetime.now(timezone.utc) + timedelta(hours=12)

        db.collection(\'rds_sessions\').document(session_id).set({
            \'uid\':       payload[\'uid\'],
            \'email\':     payload[\'email\'],
            \'name\':      payload[\'name\'],
            \'role\':      payload[\'role\'],
            \'schoolId\':  payload[\'schoolId\'],
            'createdAt': admin_firestore.FieldValue.server_timestamp(),
            \'expiresAt\': expires_at.isoformat(),
        })

        return _rds_json_resp({
            \'ok\':        True,
            \'sessionId\': session_id,
            \'user\': {
                \'uid\':      payload[\'uid\'],
                \'email\':    payload[\'email\'],
                \'name\':     payload[\'name\'],
                \'role\':     payload[\'role\'],
                \'schoolId\': payload[\'schoolId\'],
            }
        }, req=request)

    except Exception as e:
        traceback.print_exc()
        return _rds_json_resp({\'ok\': False, \'error\': str(e)}, 500, request)