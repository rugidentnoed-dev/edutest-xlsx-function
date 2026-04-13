import io
import os
import json
import sys
import traceback
import requests
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Protection, Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import make_response

import hmac
import hashlib
import base64
import uuid

import firebase_admin
from firebase_admin import credentials, firestore as admin_firestore, auth as fb_auth

# ── Firebase singleton ────────────────────────────────────────────────────────
_fdb = None

def get_db():
    """Initialise Firebase Admin once, return Firestore client."""
    global _fdb
    if _fdb:
        return _fdb
    if not firebase_admin._apps:
        sa_json = os.environ.get('SERVICE_ACCOUNT_JSON', '')
        if sa_json:
            try:
                sa_dict = json.loads(sa_json)
            except json.JSONDecodeError:
                sa_dict = json.loads(sa_json.replace('\r\n', '\n'))
            if 'private_key' in sa_dict:
                pk = sa_dict['private_key']
                if '\\n' in pk:
                    sa_dict['private_key'] = pk.replace('\\n', '\n')
            try:
                cred = credentials.Certificate(sa_dict)
            except Exception:
                traceback.print_exc()
                raise
        else:
            cred = credentials.Certificate('serviceAccountKey.json')
        try:
            firebase_admin.initialize_app(cred)
        except Exception:
            traceback.print_exc()
            raise
    try:
        _fdb = admin_firestore.client()
        return _fdb
    except Exception:
        traceback.print_exc()
        raise


# ── Environment / constants ───────────────────────────────────────────────────
PROTECT_PASSWORD = os.environ.get('XLSX_PASSWORD', 'EduTestPro2025')
RDS_URL = os.environ.get('RDS_URL', 'https://rds.edutest-pro.online')

ALLOWED_CBT_ORIGINS = {
    'https://edutest-pro.online',
    'https://edutest-pro-cbt.web.app',
    'https://edutest-pro-cbt.firebaseapp.com',
}

ALLOWED_RDS_ORIGINS = {
    'https://rds.edutest-pro.online',
    'https://edutest-rds.web.app',
    'https://edutest-rds.firebaseapp.com',
}

ALL_ALLOWED_ORIGINS = ALLOWED_CBT_ORIGINS | ALLOWED_RDS_ORIGINS


def cors_headers(req=None):
    """CORS headers for CBT-only endpoints."""
    origin  = (req.headers.get('Origin', '') if req else '') or ''
    allowed = origin if origin in ALLOWED_CBT_ORIGINS else 'https://edutest-pro.online'
    return {
        'Access-Control-Allow-Origin':  allowed,
        'Access-Control-Allow-Methods': 'POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-EduTest-Key',
        'Vary': 'Origin',
    }


def make_rds_cors(req=None):
    """CORS headers for RDS endpoints."""
    origin     = (req.headers.get('Origin', '') if req else '') or ''
    is_firebase = origin.endswith('.web.app') or origin.endswith('.firebaseapp.com')
    if origin in ALL_ALLOWED_ORIGINS or is_firebase:
        allowed = origin
    else:
        allowed = RDS_URL.rstrip('/')
    return {
        'Access-Control-Allow-Origin':      allowed,
        'Access-Control-Allow-Methods':     'POST, GET, OPTIONS',
        'Access-Control-Allow-Headers':     'Content-Type, Authorization, X-EduTest-Key',
        'Access-Control-Allow-Credentials': 'false',
        'Vary': 'Origin',
    }


CORS = cors_headers()


# ── Style constants ────────────────────────────────────────────────────────────
_GREEN       = '1B6B45'
_GREEN_LIGHT = '2D9B6A'
_DARK        = '1A1A2E'
_GREY        = '6B7280'
_WHITE       = 'FFFFFF'
_PASS_CLR    = '059669'
_FAIL_CLR    = 'DC2626'
_ALT_BG      = 'F0FDF4'

_TITLE_FILL   = PatternFill('solid', fgColor=_GREEN)
_SUB_FILL     = PatternFill('solid', fgColor=_GREEN_LIGHT)
_COL_HDR_FILL = PatternFill('solid', fgColor='166534')
_ALT_FILL     = PatternFill('solid', fgColor=_ALT_BG)
_LOCK         = Protection(locked=True)

def _font(size=10, bold=False, color=_DARK, name='Calibri'):
    return Font(name=name, size=size, bold=bold, color=color)

_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
_LEFT   = Alignment(horizontal='left',   vertical='center')
_THIN   = Side(style='thin',   color='D1FAE5')
_MED    = Side(style='medium', color='059669')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

NUM_COLS = 9


# ── Shared helpers ─────────────────────────────────────────────────────────────

def verify_token(request):
    """Verify Firebase ID token. Returns (email, uid) or raises."""
    get_db()
    hdr = request.headers.get('Authorization', '')
    if not hdr.startswith('Bearer '):
        raise PermissionError('Missing or invalid Authorization header')
    decoded = fb_auth.verify_id_token(hdr[7:])
    return decoded['email'].lower(), decoded['uid']


def secret_ok(request):
    secret = os.environ.get('XLSX_SECRET', '')
    if secret and request.headers.get('X-EduTest-Key', '') != secret:
        raise PermissionError('Unauthorized')


def json_resp(data, status=200, req=None):
    resp = make_response(json.dumps(data), status)
    resp.headers['Content-Type'] = 'application/json'
    for k, v in cors_headers(req).items():
        resp.headers[k] = v
    return resp


def rds_json_resp(data, status=200, req=None):
    resp = make_response(json.dumps(data), status)
    resp.headers['Content-Type'] = 'application/json'
    for k, v in make_rds_cors(req).items():
        resp.headers[k] = v
    return resp


def err(msg, status=400, req=None):
    return json_resp({'ok': False, 'error': msg}, status, req)


def rds_session(request, db):
    """Validate RDS session from Authorization header. Returns session dict or raises."""
    session_id = (request.headers.get('Authorization') or '').replace('Session ', '').strip()
    if not session_id:
        raise PermissionError('No session provided')
    snap = db.collection('rds_sessions').document(session_id).get()
    if not snap.exists:
        raise PermissionError('Invalid session')
    sess = snap.to_dict()
    expires_at = sess.get('expiresAt')
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
    if expires_at and datetime.now(timezone.utc) > expires_at:
        db.collection('rds_sessions').document(session_id).delete()
        raise PermissionError('Session expired')
    return sess


def ts(dt_field):
    """Convert Firestore timestamp or datetime to ISO string."""
    if dt_field and hasattr(dt_field, 'isoformat'):
        return dt_field.isoformat()
    return dt_field


def lock_all(ws):
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.protection = _LOCK
    ws.protection.set_password(PROTECT_PASSWORD)
    ws.protection.sheet               = True
    ws.protection.insertRows          = True
    ws.protection.insertColumns       = True
    ws.protection.deleteRows          = True
    ws.protection.deleteColumns       = True
    ws.protection.formatCells         = True
    ws.protection.formatRows          = True
    ws.protection.formatColumns       = True
    ws.protection.sort                = True
    ws.protection.autoFilter          = True
    ws.protection.selectLockedCells   = False
    ws.protection.selectUnlockedCells = False


def merged(ws, row, col_start, col_end, value, fill, font, align=None):
    if col_start < col_end:
        ws.merge_cells(
            start_row=row, start_column=col_start,
            end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=value)
    c.fill      = fill
    c.font      = font
    c.alignment = align or _CENTER
    return c
